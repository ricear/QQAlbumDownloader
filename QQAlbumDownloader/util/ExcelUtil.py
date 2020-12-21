# 练习：
# 封装一个ExcelUtil的模块（构造函数是excel的路径），里面提供封装的方法：
# 1 获取某个sheet对象
# 2 打印所有sheet名称
# 3 给某个sheet的某个单元格写入内容
# 4 从某个sheet的某个单元读出入内容
# 5 保存对excel对象的修改
# 6 读取某一行的内容
# 7 读取某一列的内容
from copy import copy

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import Workbook
from openpyxl.styles import colors, NamedStyle, PatternFill, Side, Border, Alignment, Protection
from openpyxl.styles import Font


class ExcelUtil():
    def __init__(self, path):
        self.path = path
        self.wb = load_workbook(path)
        self.ws = self.wb.active

    # 1 获取某个sheet对象
    def get_Excel_sheet(self, sheet_name):
        self.ws = self.wb[sheet_name]

    # 2 打印所有sheet名称
    def get_all_sheet(self):
        return self.wb.sheetnames

    # 3 给某个sheet的某个单元格写入内容
    def write_sheet(self, row1, column1, value1):
        self.ws.cell(row=row1, column=column1, value=value1)

    # 4 从某个sheet的某个单元读出内容
    def read_sheet(self, row1, column1):
        return self.ws.cell(row=row1, column=column1).value

    # 5 保存对excel对象的修改
    def save_sheet(self):
        self.wb.save(self.path)

    # 6 读取某一行的内容
    def read_row_sheet(self, row_no):
        rows = []
        for row in self.ws.iter_rows():
            rows.append(row)
        row_content = []
        for cell in rows[row_no - 1]:
            row_content.append(cell.value)
        return row_content

    # 7 读取某一列的内容
    def read_column_sheet(self, col_no):
        cols = []
        for col in self.ws.iter_cols():
            cols.append(col)
        col_content = []
        for cell in cols[col_no - 1]:
            col_content.append(cell.value)
        return col_content

    def test1(self):
         wb = Workbook()
         ws = wb.active

         data = [
             ['Apples', 10000, 5000, 8000, 6000],
             ['Pears', 2000, 3000, 4000, 5000],
             ['Bananas', 6000, 6000, 6500, 6000],
             ['Oranges', 500, 300, 200, 700],
         ]

         ws.append(["Fruit", "2011", "2012", "2013", "2014"])
         for row in data:
             ws.append(row)

         tab = Table(displayName="Table1", ref="A1:E5")

         style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                showLastColumn=False, showRowStripes=False, showColumnStripes=True)
         tab.tableStyleInfo = style
         ws.add_table(tab)

         wb.save('./s2.xlsx')

    def test2(self):
        wb = Workbook()
        ws = wb.active

        col = ws.column_dimensions['A']
        col.font = Font(bold=True)  # 将A列设定为粗体
        ws["A1"] = "99"
        row = ws.row_dimensions[1]
        row.font = Font(underline="single")  # 将第一行设定为下划线格式

        # Save the file
        wb.save('./s2.xlsx')

    # 练习：自定义字体颜色
    def test_font_color(self):
        wb = Workbook()
        ws = wb.active

        a1 = ws['A1']
        d4 = ws['D4']
        ft = Font(color=colors.GREEN)  # color="FFBB11"，颜色编码也可以设定颜色
        a1.font = ft
        d4.font = ft

        a1.font = Font(color=colors.BLUE, italic=True)
        a1.value = "abc"

        wb.save('./s2.xlsx')

    # 练习：设置某个单元格字体
    def test_cell_font(self):
        wb = Workbook()
        ws = wb.active

        a1 = ws['A1']
        d4 = ws['D4']
        a1.value = "abc"

        ft1 = Font(name=u'宋体', size=14)
        ft2 = copy(ft1)  # 复制字体对象
        ft2.name = "Tahoma"

        print(ft1.name)

        print(ft2.name)

        print(ft2.size)

        wb.save('./s2.xlsx')

    # 练习：设置单元格背景色
    def test_cell_background_color(self):

        wb = Workbook()
        ws = wb.active

        highlight = NamedStyle(name="highlight")
        highlight.font = Font(bold=True, size=20, color="ff8888")  # 设置字体颜色
        highlight.fill = PatternFill("solid", fgColor="eeE100")  # 设置背景色
        bd = Side(style='thin', color="444444")  # 设置边框thin是细，thick是粗
        highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)  # 设置边框

        print(dir(ws["A1"]))
        ws["A1"].style = highlight

        # Save the file
        wb.save('./s2.xlsx')

    # 练习：设置单元格边框和字体
    def test_cell_border_and_font(self):

        wb = Workbook()
        ws = wb.active

        ft = Font(name=u'微软雅黑',
                  size=11,
                  bold=False,
                  italic=False,
                  vertAlign=None,
                  underline='none',
                  strike=False,
                  color='FF000000')

        fill = PatternFill(fill_type="solid",
                           start_color='FFEEFFFF',
                           end_color='FF001100')

        # 边框可以选择的值为：'hair', 'medium', 'dashDot', 'dotted', 'mediumDashDot', 'dashed', 'mediumDashed', 'mediumDashDotDot', 'dashDotDot', 'slantDashDot', 'double', 'thick', 'thin']
        # diagonal 表示对角线
        bd = Border(left=Side(border_style="thin",
                              color='FF001000'),
        right = Side(border_style="thin",
                     color='FF110000'),
        top = Side(border_style="thin",
                   color='FF110000'),
        bottom = Side(border_style="thin",
                      color='FF110000'),
        diagonal = Side(border_style=None,
                        color='FF000000'),
        diagonal_direction = 0,
        outline = Side(border_style=None,
                       color='FF000000'),
        vertical = Side(border_style=None,
                        color='FF000000'),
        horizontal = Side(border_style=None,
                          color='FF110000')
        )

        alignment = Alignment(horizontal='general',

        vertical = 'bottom',
        text_rotation = 0,
        wrap_text = False,
        shrink_to_fit = False,
        indent = 0)

        number_format = 'General'

        protection = Protection(locked=True,
        hidden = False)

        ws["B5"].font = ft
        ws["B5"].fill = fill
        ws["B5"].border = bd
        ws["B5"].alignment = alignment
        ws["B5"].number_format = number_format

        ws["B5"].value = "glory road"

        # Save the file
        wb.save('./s2.xlsx')


if __name__ == "__main__":
    # Excel_obj = ExcelUtil("./s2.xlsx")
    #
    # Excel_obj.get_Excel_sheet("wangjing")
    #
    # print(Excel_obj.get_all_sheet())
    #
    # Excel_obj.write_sheet(2, 3, "10")
    #
    # print(Excel_obj.read_sheet(2, 3))
    #
    # Excel_obj.save_sheet()
    #
    # print(Excel_obj.read_row_sheet(2))
    #
    # print(Excel_obj.read_column_sheet(2))

    path = '/Users/weipeng/Personal/Projects/QQAlbumDownloader/documentations/数据字典/数据字典(玖天-外表).xlsx'
    wb = load_workbook(path)
    # for sheetname in wb.sheetnames:
    #     ws = wb[sheetname]
    #     highlight = NamedStyle(name="highlight")
    #     highlight.fill = PatternFill("solid", fgColor="00b050")  # 设置背景色
    #     for index in range(0,len(ws[1])):
    #         ws[1][index].style = highlight
    #     wb.save(path)
    wb.create_sheet('test')
    wb.save(path)
